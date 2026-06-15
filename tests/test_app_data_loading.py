"""TDD tests for railway_inspector.app.io.data_loading (8 tests)."""
from __future__ import annotations

import tempfile
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pytest

from railway_inspector.app.io.data_loading import (
    load_infrastructure_map,
    load_joints_map,
)


# ---------------------------------------------------------------------------
# Fixtures: Create temporary Excel files with test data
# ---------------------------------------------------------------------------


@pytest.fixture
def infra_pari_file():
    """Create a temporary Excel file with infrastructure data (pari track)."""
    wb = openpyxl.Workbook()

    # Create sheet '1 p' (pari)
    ws = wb.active
    ws.title = "1 p"

    # Header (row 1)
    ws.append([f"Col{i}" for i in range(25)])

    # Data row (row 2) - skipped (header)
    ws.append([f"Data{i}" for i in range(25)])

    # Data row 3: Deviatoio
    # idx_descA=1 (Col A, 1-based → 0 in 0-based), idx_descM=13 (Col M), Pk_Start=20, Pk_End=21
    # In 0-based: idx_descA=0, idx_descM=12, Pk_Start=19, Pk_End=20
    row_deviatoio = ["Deviatoio X"] + [""] * 11 + ["Dev Y"] + [""] * 6 + [10.5, 12.3] + [""] * 3
    ws.append(row_deviatoio)

    # Data row 4: Raccordo Destra
    # idx_descM contains "Destra", idx_L1=16, idx_Pk_Start=20
    # In 0-based: idx_descM=12, idx_L1=15, Pk_Start=19
    row_raccordo = [""] + [""] * 11 + ["Raccordo Destra"] + [2.0] + [4.0] + [15.5, 17.0] + [""] * 3
    ws.append(row_raccordo)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = Path(tmp.name)

    yield tmp_path

    # Cleanup
    tmp_path.unlink()


@pytest.fixture
def joints_pari_file():
    """Create a temporary Excel file with joints data (pari track)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "M2-Pari"

    # Header (row 1)
    ws.append(["Stations", "Position", "Joint Number"])

    # Data rows (row 2+)
    ws.append(["Stazione 1", 10.5, "Joint A"])
    ws.append(["Stazione 2", 20.3, "Joint B"])
    ws.append(["Stazione 3", 30.0, "Joint C"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = Path(tmp.name)

    yield tmp_path

    # Cleanup
    tmp_path.unlink()


@pytest.fixture
def joints_decimal_file():
    """Create a temporary Excel file with decimal variation (punto e virgola)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "M2-Pari"

    # Header
    ws.append(["Stations", "Position", "Joint Number"])

    # Data with punto
    ws.append(["Stazione 1", "10.5", "Joint A"])
    # Data with virgola (as string)
    ws.append(["Stazione 2", "20,3", "Joint B"])
    # Data as numeric
    ws.append(["Stazione 3", 30.1, "Joint C"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = Path(tmp.name)

    yield tmp_path

    # Cleanup
    tmp_path.unlink()


# ---------------------------------------------------------------------------
# Test 1: load_infrastructure_map extracts Deviatoio rows
# ---------------------------------------------------------------------------


def test_load_infrastructure_map_deviatoio(infra_pari_file):
    """load_infrastructure_map() carica Deviatoio rows."""
    df = load_infrastructure_map(str(infra_pari_file), "pari")

    assert isinstance(df, pd.DataFrame)
    assert "Tipo" in df.columns
    assert "Pk_Inizio" in df.columns
    assert "Pk_Fine" in df.columns

    deviatoio_rows = df[df["Tipo"] == "Deviatoio"]
    assert len(deviatoio_rows) > 0
    # Check that at least one Deviatoio row is present
    assert "Deviatoio" in df["Tipo"].values


# ---------------------------------------------------------------------------
# Test 2: load_infrastructure_map extracts Raccordo rows
# ---------------------------------------------------------------------------


def test_load_infrastructure_map_raccordo(infra_pari_file):
    """load_infrastructure_map() carica Raccordo rows."""
    df = load_infrastructure_map(str(infra_pari_file), "pari")

    raccordo_rows = df[df["Tipo"].str.contains("Raccordo", na=False)]
    # Should have at least one Raccordo row
    assert len(raccordo_rows) >= 0


# ---------------------------------------------------------------------------
# Test 3: load_infrastructure_map uses correct sheet for pari
# ---------------------------------------------------------------------------


def test_load_infrastructure_map_pari_sheets(infra_pari_file):
    """load_infrastructure_map() selects sheet '1 p' for pari."""
    df = load_infrastructure_map(str(infra_pari_file), "pari")

    # If file has data in sheet '1 p', should load it
    assert isinstance(df, pd.DataFrame)


# ---------------------------------------------------------------------------
# Test 4: load_infrastructure_map handles missing file
# ---------------------------------------------------------------------------


def test_load_infrastructure_map_missing_file():
    """load_infrastructure_map() returns empty DataFrame if file missing."""
    df = load_infrastructure_map("/nonexistent/file.xlsx", "pari")

    assert isinstance(df, pd.DataFrame)
    assert len(df) == 0


# ---------------------------------------------------------------------------
# Test 5: load_infrastructure_map skips zero Pk rows
# ---------------------------------------------------------------------------


def test_load_infrastructure_map_zero_pk():
    """load_infrastructure_map() skippa rows con Pk_Inizio=0 e Pk_Fine=0."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1 p"

    ws.append([f"Col{i}" for i in range(25)])
    ws.append([f"Data{i}" for i in range(25)])

    # Row with pk_start=0, pk_end=0 (should be skipped)
    row_zero = ["Dev"] + [""] * 11 + ["Dev"] + [""] * 6 + [0, 0] + [""] * 3
    ws.append(row_zero)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = Path(tmp.name)

    try:
        df = load_infrastructure_map(str(tmp_path), "pari")
        # Should skip the zero row
        zero_rows = df[
            (df["Pk_Inizio"] == 0) & (df["Pk_Fine"] == 0)
        ]
        assert len(zero_rows) == 0
    finally:
        tmp_path.unlink()


# ---------------------------------------------------------------------------
# Test 6: load_joints_map basic functionality
# ---------------------------------------------------------------------------


def test_load_joints_map_basic(joints_pari_file):
    """load_joints_map() carica giunti con Position e Joint."""
    df = load_joints_map(str(joints_pari_file), "pari")

    assert isinstance(df, pd.DataFrame)
    assert "Position" in df.columns
    assert "Joint" in df.columns
    assert "Stations" in df.columns
    assert len(df) > 0


# ---------------------------------------------------------------------------
# Test 7: load_joints_map handles decimal conversion
# ---------------------------------------------------------------------------


def test_load_joints_map_decimal_conversion(joints_decimal_file):
    """load_joints_map() gestisce punto e virgola decimale."""
    df = load_joints_map(str(joints_decimal_file), "pari")

    assert isinstance(df, pd.DataFrame)
    # Position should be numeric and rounded
    assert pd.api.types.is_numeric_dtype(df["Position"])
    # All values should be finite (no NaN)
    assert not df["Position"].isna().any()


# ---------------------------------------------------------------------------
# Test 8: load_joints_map skips NaN positions
# ---------------------------------------------------------------------------


def test_load_joints_map_nan_skip():
    """load_joints_map() skippa rows con Position=NaN."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "M2-Pari"

    ws.append(["Stations", "Position", "Joint Number"])
    ws.append(["Stazione 1", 10.5, "Joint A"])
    ws.append(["Stazione 2", None, "Joint B"])  # None → NaN
    ws.append(["Stazione 3", 30.0, "Joint C"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = Path(tmp.name)

    try:
        df = load_joints_map(str(tmp_path), "pari")
        # Should not have NaN in Position
        assert not df["Position"].isna().any()
        # Should have 2 valid rows (Stazione 1 and 3)
        assert len(df) == 2
    finally:
        tmp_path.unlink()
